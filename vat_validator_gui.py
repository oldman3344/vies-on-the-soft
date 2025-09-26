#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VAT验证工具 - PyQt5 GUI版本
支持单个验证、批量验证、Excel导入导出功能
"""

import sys
import os
import requests
import json
import time
import shutil
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QFileDialog,
    QProgressBar, QTabWidget, QTableWidget, QTableWidgetItem,
    QMessageBox, QComboBox, QGroupBox, QGridLayout, QSplitter,
    QHeaderView, QStatusBar, QFrame, QCheckBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QPropertyAnimation, QRect
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor, QPainter, QPen, QMovie
from openpyxl import load_workbook, Workbook
from document_processor import DocumentProcessor, create_default_column_mapping



class VATValidationWorker(QThread):
    """
    VAT验证工作线程 - 支持并发处理
    """
    progress_updated = pyqtSignal(int)
    result_ready = pyqtSignal(dict)
    finished_all = pyqtSignal(list)
    log_message = pyqtSignal(str)
    
    def __init__(self, vat_data, parent=None):
        super().__init__(parent)
        self.vat_data = vat_data
        self.results = []
        self.is_running = True
        self.completed_count = 0
        self.lock = Lock()
        self.session = requests.Session()
        self.cache = {}  # 简单的内存缓存
        self.cache_lock = Lock()  # 缓存锁
        self.max_workers = 4  # 并发线程数 - 降低到4个
        self.request_delay = 0.25  # 请求间隔（秒）- 增加到0.25秒
        self.max_retries = 3  # 最大重试次数
        self.retry_delay = 0.5  # 重试间隔（秒）
    
    def stop(self):
        self.is_running = False
        self.session.close()
    
    def extract_country_code(self, vat_number):
        """
        从VAT号码中提取国家代码
        """
        eu_countries = {
            'AT', 'BE', 'BG', 'CY', 'CZ', 'DE', 'DK', 'EE', 'ES', 'FI',
            'FR', 'GR', 'HR', 'HU', 'IE', 'IT', 'LT', 'LU', 'LV', 'MT',
            'NL', 'PL', 'PT', 'RO', 'SE', 'SI', 'SK'
        }
        
        if len(vat_number) >= 2:
            potential_code = vat_number[:2].upper()
            if potential_code in eu_countries:
                return potential_code
        return None
    
    def verify_vat(self, country_code, vat_number, original_data=None):
        """
        验证单个VAT号码 - 支持缓存、连接复用和智能重试
        """
        # vat_number 应该是完整的VAT号码（包含国家代码前缀）
        full_vat = vat_number if vat_number.startswith(country_code) else f"{country_code}{vat_number}"
        
        # 检查缓存（只返回成功的缓存结果）
        with self.cache_lock:
            if full_vat in self.cache and self.cache[full_vat]['success']:
                result = self.cache[full_vat].copy()
                if original_data:
                    result['original_data'] = original_data
                return result
        
        # 初始化result变量
        result = {
            'success': False,
            'vat_number': full_vat,
            'country_code': country_code,
            'error': '未知错误',
            'attempts': 0
        }
        
        # 执行带重试的验证
        for attempt in range(self.max_retries):
            try:
                url = f"https://ec.europa.eu/taxation_customs/vies/rest-api/ms/{country_code}/vat/{full_vat}"
                headers = {
                    'Accept': 'application/json, text/plain, */*',
                    'Accept-Encoding': 'gzip, deflate, br, zstd',
                    'Accept-Language': 'zh-CN,zh;q=0.9',
                    'Cache-Control': 'No-Cache',
                    'Connection': 'keep-alive',
                    'Host': 'ec.europa.eu',
                    'Pragma': 'no-cache',
                    'Referer': 'https://ec.europa.eu/taxation_customs/vies/',
                    'Sec-CH-UA': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
                    'Sec-CH-UA-Mobile': '?0',
                    'Sec-CH-UA-Platform': '"macOS"',
                    'Sec-Fetch-Dest': 'empty',
                    'Sec-Fetch-Mode': 'cors',
                    'Sec-Fetch-Site': 'same-origin',
                    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
                    'X-Requested-With': 'XMLHttpRequest'
                }
                
                # 发送请求日志到GUI
                log_msg = f"\n=== VAT验证请求 (尝试 {attempt + 1}/{self.max_retries}) ===\n"
                log_msg += f"请求URL: {url}\n"
                log_msg += f"请求头: {json.dumps(headers, indent=2, ensure_ascii=False)}\n"
                self.log_message.emit(log_msg)
                
                # 添加请求间隔
                time.sleep(self.request_delay)
                
                response = self.session.get(url, headers=headers, timeout=15)
                
                # 发送响应日志到GUI
                response_log = f"响应状态码: {response.status_code}\n"
                response_log += f"响应头: {dict(response.headers)}\n"
                if response.status_code == 200:
                    try:
                        response_data = response.json()
                        response_log += f"响应数据: {json.dumps(response_data, indent=2, ensure_ascii=False)}\n"
                    except:
                        response_log += f"响应内容: {response.text[:500]}...\n"
                self.log_message.emit(response_log)
                
                if response.status_code == 200:
                    data = response.json()
                    # 验证响应数据的完整性
                    if 'isValid' in data and 'requestDate' in data:
                        result = {
                            'success': True,
                            'vat_number': full_vat,
                            'country_code': country_code,
                            'data': data,
                            'attempts': attempt + 1
                        }
                        # 只缓存成功的结果
                        with self.cache_lock:
                            self.cache[full_vat] = result.copy()
                        break
                    else:
                        # 数据不完整，继续重试
                        if attempt < self.max_retries - 1:
                            time.sleep(self.retry_delay)
                            continue
                        else:
                            result = {
                                'success': False,
                                'vat_number': full_vat,
                                'country_code': country_code,
                                'error': f"响应数据不完整，重试{self.max_retries}次后失败",
                                'attempts': attempt + 1
                            }
                elif response.status_code in [429, 503, 504]:  # 限流或服务器错误，重试
                    if attempt < self.max_retries - 1:
                        time.sleep(self.retry_delay * (attempt + 1))  # 指数退避
                        continue
                    else:
                        result = {
                            'success': False,
                            'vat_number': full_vat,
                            'country_code': country_code,
                            'error': f"服务器错误 {response.status_code}，重试{self.max_retries}次后失败",
                            'attempts': attempt + 1
                        }
                else:
                    result = {
                        'success': False,
                        'vat_number': full_vat,
                        'country_code': country_code,
                        'error': f"HTTP错误: {response.status_code}",
                        'attempts': attempt + 1
                    }
                    break  # 对于其他HTTP错误，不重试
                    
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                if attempt < self.max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                    continue
                else:
                    result = {
                        'success': False,
                        'vat_number': full_vat,
                        'country_code': country_code,
                        'error': f"网络错误: {str(e)}，重试{self.max_retries}次后失败",
                        'attempts': attempt + 1
                    }
            except Exception as e:
                result = {
                    'success': False,
                    'vat_number': full_vat,
                    'country_code': country_code,
                    'error': f"验证失败: {str(e)}",
                    'attempts': attempt + 1
                }
                break  # 对于其他异常，不重试
        
        if original_data:
            result['original_data'] = original_data
        
        return result
    
    def process_single_vat(self, vat_item):
        """
        处理单个VAT验证任务
        """
        vat_number, original_data = vat_item
        
        if not self.is_running:
            return None
            
        # 检查是否有预设的国家代码
        if 'country_code' in original_data:
            country_code = original_data['country_code']
            # 保持完整的VAT号码
            full_vat = vat_number
        else:
            # 提取国家代码
            country_code = self.extract_country_code(vat_number)
            
            if country_code:
                # 保持完整的VAT号码（包含国家代码前缀）
                full_vat = vat_number
            else:
                # 无法识别国家代码，返回错误结果
                return {
                    'success': False,
                    'vat_number': vat_number,
                    'country_code': 'UNKNOWN',
                    'error': 'Unable to extract valid EU country code from VAT number',
                    'attempts': 0
                }
        
        result = self.verify_vat(country_code, full_vat, original_data)
        return result
    
    def run(self):
        total = len(self.vat_data)
        
        # 使用线程池并发处理
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 提交所有任务
            future_to_index = {}
            for i, vat_item in enumerate(self.vat_data):
                if not self.is_running:
                    break
                future = executor.submit(self.process_single_vat, vat_item)
                future_to_index[future] = i
            
            # 处理完成的任务
            for future in as_completed(future_to_index):
                if not self.is_running:
                    break
                    
                try:
                    result = future.result()
                    if result is not None:
                        with self.lock:
                            self.results.append(result)
                            self.completed_count += 1
                            
                        # 发送进度和结果信号
                        progress = int(self.completed_count / total * 100)
                        self.progress_updated.emit(progress)
                        self.result_ready.emit(result)
                        
                except Exception as e:
                    # 处理异常情况
                    error_result = {
                        'success': False,
                        'vat_number': 'Unknown',
                        'country_code': 'Unknown',
                        'error': f"处理异常: {str(e)}",
                        'original_data': {}
                    }
                    with self.lock:
                        self.results.append(error_result)
                        self.completed_count += 1
                    
                    progress = int(self.completed_count / total * 100)
                    self.progress_updated.emit(progress)
                    self.result_ready.emit(error_result)
        
        if self.is_running:
            self.finished_all.emit(self.results)

class VATValidatorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.results = []
        self.worker = None
        
        try:
            print("正在初始化用户界面...")
            self.init_ui()
            print("✓ 用户界面初始化完成")
        except Exception as e:
            print(f"✗ 用户界面初始化失败: {e}")
            import traceback
            traceback.print_exc()
            raise
        
    def init_ui(self):
        """
        初始化用户界面
        """
        self.setWindowTitle("VAT验证工具 v2.0")
        self.setGeometry(100, 100, 1200, 800)
        
        # 设置应用样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTabWidget::pane {
                border: 1px solid #c0c0c0;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #e1e1e1;
                padding: 8px 16px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: white;
                border-bottom: 2px solid #0078d4;
            }
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106ebe;
            }
            QPushButton:pressed {
                background-color: #005a9e;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            QLineEdit {
                padding: 8px;
                border: 2px solid #e1e1e1;
                border-radius: 4px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border-color: #0078d4;
            }
            QTextEdit {
                border: 2px solid #e1e1e1;
                border-radius: 4px;
                font-family: 'Courier New', monospace;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #e1e1e1;
                border-radius: 4px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        # 创建中央部件和标签页
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout(central_widget)
        
        # 创建标签页
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        
        # 单个验证标签页
        print("正在创建单个验证标签页...")
        self.create_single_validation_tab()
        print("✓ 单个验证标签页创建完成")
        
        # 批量验证标签页
        print("正在创建批量验证标签页...")
        self.create_batch_validation_tab()
        print("✓ 批量验证标签页创建完成")
        
        # 结果查看标签页
        print("正在创建结果查看标签页...")
        self.create_results_tab()
        print("✓ 结果查看标签页创建完成")
        
        # 请求日志标签页
        print("正在创建请求日志标签页...")
        self.create_log_tab()
        print("✓ 请求日志标签页创建完成")
        
        # 文档处理标签页
        print("正在创建文档处理标签页...")
        self.create_document_processing_tab()
        print("✓ 文档处理标签页创建完成")
        
        # 状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")
        
    def create_single_validation_tab(self):
        """
        创建单个验证标签页
        """
        tab = QWidget()
        self.tab_widget.addTab(tab, "单个验证")
        
        layout = QVBoxLayout(tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 输入组
        input_group = QGroupBox("VAT号码输入")
        input_layout = QGridLayout(input_group)
        
        # VAT号码输入
        input_layout.addWidget(QLabel("VAT号码:"), 0, 0)
        self.vat_input = QLineEdit()
        self.vat_input.setPlaceholderText("例如: IT05159640266 或 05159640266")
        input_layout.addWidget(self.vat_input, 0, 1)
        
        # 国家代码选择
        input_layout.addWidget(QLabel("国家代码:"), 1, 0)
        self.country_combo = QComboBox()
        self.country_combo.addItems([
            "自动检测", "AT", "BE", "BG", "CY", "CZ", "DE", "DK", "EE", "ES", "FI",
            "FR", "GR", "HR", "HU", "IE", "IT", "LT", "LU", "LV", "MT",
            "NL", "PL", "PT", "RO", "SE", "SI", "SK"
        ])
        input_layout.addWidget(self.country_combo, 1, 1)
        
        # 验证按钮
        self.verify_button = QPushButton("验证VAT号码")
        self.verify_button.clicked.connect(self.verify_single_vat)
        input_layout.addWidget(self.verify_button, 2, 0, 1, 2)
        
        layout.addWidget(input_group)
        
        # 结果显示组
        result_group = QGroupBox("验证结果")
        result_layout = QVBoxLayout(result_group)
        
        self.single_result_text = QTextEdit()
        self.single_result_text.setMaximumHeight(300)
        result_layout.addWidget(self.single_result_text)
        
        layout.addWidget(result_group)
        
        # 添加弹性空间
        layout.addStretch()
        
    def create_batch_validation_tab(self):
        """
        创建批量验证标签页
        """
        tab = QWidget()
        self.tab_widget.addTab(tab, "批量验证")
        
        layout = QVBoxLayout(tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 文件操作组
        file_group = QGroupBox("Excel文件操作")
        file_layout = QHBoxLayout(file_group)
        
        self.file_path_label = QLabel("未选择文件")
        file_layout.addWidget(self.file_path_label)
        
        self.browse_button = QPushButton("选择Excel文件")
        self.browse_button.clicked.connect(self.browse_excel_file)
        file_layout.addWidget(self.browse_button)
        
        self.start_batch_button = QPushButton("开始批量验证")
        self.start_batch_button.clicked.connect(self.start_batch_validation)
        self.start_batch_button.setEnabled(False)
        file_layout.addWidget(self.start_batch_button)
        
        self.stop_batch_button = QPushButton("停止验证")
        self.stop_batch_button.clicked.connect(self.stop_batch_validation)
        self.stop_batch_button.setEnabled(False)
        file_layout.addWidget(self.stop_batch_button)
        
        layout.addWidget(file_group)
        
        # 进度组
        progress_group = QGroupBox("验证进度")
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        progress_layout.addWidget(self.progress_bar)
        
        self.progress_label = QLabel("等待开始...")
        progress_layout.addWidget(self.progress_label)
        
        layout.addWidget(progress_group)
        
        # 实时结果显示
        realtime_group = QGroupBox("实时验证结果")
        realtime_layout = QVBoxLayout(realtime_group)
        
        self.batch_result_text = QTextEdit()
        realtime_layout.addWidget(self.batch_result_text)
        
        layout.addWidget(realtime_group)
        
    def create_results_tab(self):
        """
        创建结果查看标签页
        """
        tab = QWidget()
        self.tab_widget.addTab(tab, "结果查看")
        
        layout = QVBoxLayout(tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 搜索功能组
        search_group = QGroupBox("搜索")
        search_layout = QHBoxLayout(search_group)
        
        search_layout.addWidget(QLabel("搜索:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入VAT号码、公司名称或其他信息进行搜索...")
        self.search_input.textChanged.connect(self.filter_results)
        search_layout.addWidget(self.search_input)
        
        self.clear_search_button = QPushButton("清空搜索")
        self.clear_search_button.clicked.connect(self.clear_search)
        search_layout.addWidget(self.clear_search_button)
        
        layout.addWidget(search_group)
        
        # 操作按钮组
        button_group = QGroupBox("操作")
        button_layout = QHBoxLayout(button_group)
        
        self.export_button = QPushButton("导出结果到Excel")
        self.export_button.clicked.connect(self.export_results)
        self.export_button.setEnabled(False)
        button_layout.addWidget(self.export_button)
        
        self.clear_button = QPushButton("清空结果")
        self.clear_button.clicked.connect(self.clear_results)
        button_layout.addWidget(self.clear_button)
        
        button_layout.addStretch()
        
        layout.addWidget(button_group)
        
        # 结果表格
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(8)
        self.results_table.setHorizontalHeaderLabels([
            "VAT号码", "国家代码", "验证状态", "是否有效", "公司名称", "地址", "验证时间", "错误信息"
        ])
        
        # 设置表格属性
        header = self.results_table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
            header.setSectionResizeMode(QHeaderView.Interactive)
        
        layout.addWidget(self.results_table)
        
    def verify_single_vat(self):
        """
        验证单个VAT号码
        """
        vat_number = self.vat_input.text().strip()
        if not vat_number:
            QMessageBox.warning(self, "警告", "请输入VAT号码")
            return
        
        self.verify_button.setEnabled(False)
        self.single_result_text.clear()
        self.single_result_text.append("正在验证...")
        self.status_bar.showMessage("正在验证VAT号码...")
        
        # 获取国家代码
        country_selection = self.country_combo.currentText()
        if country_selection == "自动检测":
            country_code = self.extract_country_code(vat_number)
            if country_code:
                clean_vat = vat_number[2:] if vat_number.startswith(country_code) else vat_number
            else:
                country_code = "IT"  # 默认意大利
                clean_vat = vat_number
        else:
            country_code = country_selection
            clean_vat = vat_number
        
        # 创建工作线程，传递国家代码
        self.single_worker = VATValidationWorker([(clean_vat, {'country_code': country_code})])
        self.single_worker.result_ready.connect(self.handle_single_result)
        self.single_worker.finished.connect(lambda: self.verify_button.setEnabled(True))
        self.single_worker.log_message.connect(self.append_log)
        self.single_worker.start()
    
    def handle_single_result(self, result):
        """
        处理单个验证结果
        """
        self.single_result_text.clear()
        
        if result['success']:
            data = result['data']
            is_valid = data.get('isValid', False)
            
            result_text = f"VAT号码: {result['vat_number']}\n"
            result_text += f"国家代码: {result['country_code']}\n"
            result_text += f"验证状态: 成功\n"
            result_text += f"是否有效: {'是' if is_valid else '否'}\n"
            
            if is_valid:
                result_text += f"公司名称: {data.get('name', '未提供')}\n"
                result_text += f"地址: {data.get('address', '未提供')}\n"
            
            result_text += f"验证时间: {data.get('requestDate', '未提供')}\n"
            
            self.single_result_text.append(result_text)
            
            if is_valid:
                self.status_bar.showMessage(f"验证成功: {result['vat_number']} 有效")
            else:
                self.status_bar.showMessage(f"验证成功: {result['vat_number']} 无效")
        else:
            error_text = f"VAT号码: {result['vat_number']}\n"
            error_text += f"验证状态: 失败\n"
            error_text += f"错误信息: {result['error']}\n"
            
            self.single_result_text.append(error_text)
            self.status_bar.showMessage(f"验证失败: {result['error']}")
    
    def extract_country_code(self, vat_number):
        """
        从VAT号码中提取国家代码
        """
        eu_countries = {
            'AT', 'BE', 'BG', 'CY', 'CZ', 'DE', 'DK', 'EE', 'ES', 'FI',
            'FR', 'GR', 'HR', 'HU', 'IE', 'IT', 'LT', 'LU', 'LV', 'MT',
            'NL', 'PL', 'PT', 'RO', 'SE', 'SI', 'SK'
        }
        
        if len(vat_number) >= 2:
            potential_code = vat_number[:2].upper()
            if potential_code in eu_countries:
                return potential_code
        return None
    
    def browse_excel_file(self):
        """
        浏览Excel文件
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path_label.setText(os.path.basename(file_path))
            self.excel_file_path = file_path
            self.start_batch_button.setEnabled(True)
    
    def start_batch_validation(self):
        """
        开始批量验证
        """
        if not hasattr(self, 'excel_file_path'):
            QMessageBox.warning(self, "警告", "请先选择Excel文件")
            return
        
        # 读取Excel文件
        try:
            vat_data = self.load_excel_data(self.excel_file_path)
            if not vat_data:
                return
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取Excel文件失败: {str(e)}")
            return
        
        # 清空之前的结果
        self.results = []
        self.batch_result_text.clear()
        self.results_table.setRowCount(0)
        
        # 设置UI状态
        self.start_batch_button.setEnabled(False)
        self.stop_batch_button.setEnabled(True)
        self.export_button.setEnabled(False)
        
        # 重置进度条
        self.progress_bar.setValue(0)
        self.progress_label.setText(f"开始验证 {len(vat_data)} 个VAT号码...")
        
        # 创建并启动工作线程
        self.worker = VATValidationWorker(vat_data)
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.result_ready.connect(self.handle_batch_result)
        self.worker.finished_all.connect(self.batch_validation_finished)
        self.worker.log_message.connect(self.append_log)
        self.worker.start()
        
        self.status_bar.showMessage("批量验证进行中...")
    
    def load_excel_data(self, file_path):
        """
        加载Excel数据
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            if ws is None:
                QMessageBox.critical(self, "错误", "无法读取Excel工作表")
                return []
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value is not None else '')
            
            # 检查必要的列
            if 'NIF Contraparte' not in headers:
                QMessageBox.critical(self, "错误", "Excel文件必须包含'NIF Contraparte'列")
                return []
            
            # 找到NIF列的索引
            nif_col_index = headers.index('NIF Contraparte')
            
            # 提取数据
            vat_data = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if row and len(row) > nif_col_index and row[nif_col_index]:
                    vat_number = str(row[nif_col_index]).strip()
                    if vat_number:
                        row_data = {}
                        for i in range(min(len(headers), len(row))):
                            row_data[headers[i]] = row[i]
                        row_data['_row_number'] = row_num
                        vat_data.append((vat_number, row_data))
            
            if not vat_data:
                QMessageBox.warning(self, "警告", "未找到有效的VAT号码")
                return []
            
            return vat_data
            
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")
    
    def update_progress(self, value):
        """
        更新进度条
        """
        self.progress_bar.setValue(value)
        self.progress_label.setText(f"验证进度: {value}%")
    
    def handle_batch_result(self, result):
        """
        处理批量验证结果
        """
        self.results.append(result)
        
        # 添加到实时显示
        if result['success']:
            data = result['data']
            is_valid = data.get('isValid', False)
            company_name = data.get('name', '未提供')
            status_text = f"✅ {result['vat_number']}: {'有效' if is_valid else '无效'}"
            if is_valid:
                status_text += f" - {company_name}"
        else:
            status_text = f"❌ {result['vat_number']}: 验证失败 - {result['error']}"
        
        self.batch_result_text.append(status_text)
        
        # 添加到结果表格
        self.add_result_to_table(result)
    
    def add_result_to_table(self, result):
        """
        添加结果到表格
        """
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)
        
        self.results_table.setItem(row, 0, QTableWidgetItem(result['vat_number']))
        self.results_table.setItem(row, 1, QTableWidgetItem(result['country_code']))
        self.results_table.setItem(row, 2, QTableWidgetItem('成功' if result['success'] else '失败'))
        
        if result['success'] and 'data' in result:
            data = result['data']
            self.results_table.setItem(row, 3, QTableWidgetItem('是' if data.get('isValid', False) else '否'))
            self.results_table.setItem(row, 4, QTableWidgetItem(data.get('name', '')))
            self.results_table.setItem(row, 5, QTableWidgetItem(data.get('address', '')))
            self.results_table.setItem(row, 6, QTableWidgetItem(data.get('requestDate', '')))
            self.results_table.setItem(row, 7, QTableWidgetItem(''))
        else:
            self.results_table.setItem(row, 3, QTableWidgetItem(''))
            self.results_table.setItem(row, 4, QTableWidgetItem(''))
            self.results_table.setItem(row, 5, QTableWidgetItem(''))
            self.results_table.setItem(row, 6, QTableWidgetItem(''))
            self.results_table.setItem(row, 7, QTableWidgetItem(result.get('error', '')))
    
    def batch_validation_finished(self, results):
        """
        批量验证完成
        """
        self.start_batch_button.setEnabled(True)
        self.stop_batch_button.setEnabled(False)
        self.export_button.setEnabled(True)
        
        # 统计结果
        total = len(results)
        successful = sum(1 for r in results if r['success'])
        valid = sum(1 for r in results if r['success'] and r.get('data', {}).get('isValid', False))
        
        self.progress_label.setText(f"验证完成: 总数{total}, 成功{successful}, 有效{valid}")
        self.status_bar.showMessage(f"批量验证完成: {total}个VAT号码, {successful}个验证成功, {valid}个有效")
        
        # 显示完成消息
        QMessageBox.information(
            self, "验证完成", 
            f"批量验证已完成!\n\n总数: {total}\n验证成功: {successful}\n有效VAT: {valid}"
        )
    
    def stop_batch_validation(self):
        """
        停止批量验证
        """
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait()
            
            self.start_batch_button.setEnabled(True)
            self.stop_batch_button.setEnabled(False)
            self.export_button.setEnabled(True)
            
            self.progress_label.setText("验证已停止")
            self.status_bar.showMessage("批量验证已停止")
    
    def export_results(self):
        """
        导出结果到Excel
        """
        if not self.results:
            QMessageBox.warning(self, "警告", "没有验证结果可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存验证结果", f"vat_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel文件 (*.xlsx)"
        )
        
        if file_path:
            try:
                self.save_results_to_excel(file_path)
                QMessageBox.information(self, "成功", f"结果已成功导出到:\n{file_path}")
                self.status_bar.showMessage(f"结果已导出到: {os.path.basename(file_path)}")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def save_results_to_excel(self, file_path):
        """
        保存结果到Excel文件
        """
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "VAT验证结果"
        
        # 设置表头 - 按照用户要求的格式
        headers = ['NIF Contraparte', 'Importe', 'Tipo', 'Name', 'IS_Valid']
        
        # 写入表头
        if ws is not None:
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row_idx, result in enumerate(self.results, 2):
                # NIF Contraparte - VAT号码
                ws.cell(row=row_idx, column=1, value=result['vat_number'])
                
                # Importe - 从原始数据中获取，如果没有则为空
                importe = ''
                if 'original_data' in result and 'Importe' in result['original_data']:
                    importe = result['original_data']['Importe']
                ws.cell(row=row_idx, column=2, value=importe)
                
                # Tipo - 从原始数据中获取，如果没有则为空
                tipo = ''
                if 'original_data' in result and 'Tipo' in result['original_data']:
                    tipo = result['original_data']['Tipo']
                ws.cell(row=row_idx, column=3, value=tipo)
                
                # Name - 公司名称
                name = ''
                if result['success'] and 'data' in result:
                    name = result['data'].get('name', '')
                ws.cell(row=row_idx, column=4, value=name)
                
                # IS_Valid - 是否有效
                is_valid = 'False'
                if result['success'] and 'data' in result:
                    is_valid = 'True' if result['data'].get('isValid', False) else 'False'
                ws.cell(row=row_idx, column=5, value=is_valid)
        
        wb.save(file_path)
    
    def clear_results(self):
        """
        清空结果
        """
        reply = QMessageBox.question(
            self, "确认", "确定要清空所有验证结果吗?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.results = []
            self.batch_result_text.clear()
            self.results_table.setRowCount(0)
            self.export_button.setEnabled(False)
            self.progress_bar.setValue(0)
            self.progress_label.setText("等待开始...")
            self.status_bar.showMessage("结果已清空")
    
    def create_log_tab(self):
        """
        创建请求日志标签页
        """
        tab = QWidget()
        self.tab_widget.addTab(tab, "请求日志")
        
        layout = QVBoxLayout(tab)
        
        # 日志显示区域
        log_group = QGroupBox("API请求日志")
        log_layout = QVBoxLayout(log_group)
        
        # 日志文本框
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Courier New", 10))
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #ffffff;
                border: 1px solid #555;
                font-family: 'Courier New', monospace;
            }
        """)
        log_layout.addWidget(self.log_text)
        
        # 控制按钮
        button_layout = QHBoxLayout()
        
        self.clear_log_button = QPushButton("清空日志")
        self.clear_log_button.clicked.connect(self.clear_log)
        button_layout.addWidget(self.clear_log_button)
        
        self.export_log_button = QPushButton("导出日志")
        self.export_log_button.clicked.connect(self.export_log)
        button_layout.addWidget(self.export_log_button)
        
        self.auto_scroll_checkbox = QCheckBox("自动滚动")
        self.auto_scroll_checkbox.setChecked(True)
        button_layout.addWidget(self.auto_scroll_checkbox)
        
        button_layout.addStretch()
        
        log_layout.addLayout(button_layout)
        layout.addWidget(log_group)
    
    def append_log(self, message):
        """
        添加日志消息
        """
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        
        self.log_text.append(formatted_message)
        
        # 自动滚动到底部
        if hasattr(self, 'auto_scroll_checkbox') and self.auto_scroll_checkbox.isChecked():
            cursor = self.log_text.textCursor()
            cursor.movePosition(cursor.End)
            self.log_text.setTextCursor(cursor)
    
    def clear_log(self):
        """
        清空日志
        """
        self.log_text.clear()
        self.append_log("日志已清空")
    
    def export_log(self):
        """
        导出日志到txt文件
        """
        if not self.log_text.toPlainText().strip():
            QMessageBox.warning(self, "警告", "没有日志内容可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存请求日志", f"vat_request_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            "文本文件 (*.txt)"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(self.log_text.toPlainText())
                QMessageBox.information(self, "成功", f"日志已成功导出到:\n{file_path}")
                self.status_bar.showMessage(f"日志已导出到: {os.path.basename(file_path)}")
            except Exception as e:
                 QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def filter_results(self):
        """
        根据搜索条件过滤结果表格
        """
        search_text = self.search_input.text().lower().strip()
        
        for row in range(self.results_table.rowCount()):
            should_show = True
            
            if search_text:
                # 检查所有列是否包含搜索文本
                row_text = ""
                for col in range(self.results_table.columnCount()):
                    item = self.results_table.item(row, col)
                    if item:
                        row_text += item.text().lower() + " "
                
                should_show = search_text in row_text
            
            self.results_table.setRowHidden(row, not should_show)
        
        # 更新状态栏显示过滤结果
        if search_text:
            visible_count = sum(1 for row in range(self.results_table.rowCount()) 
                              if not self.results_table.isRowHidden(row))
            total_count = self.results_table.rowCount()
            self.status_bar.showMessage(f"搜索结果: {visible_count}/{total_count} 条记录")
        else:
            self.status_bar.showMessage(f"显示全部 {self.results_table.rowCount()} 条记录")
    
    def clear_search(self):
        """
        清空搜索条件
        """
        self.search_input.clear()
        # filter_results会通过textChanged信号自动调用
    
    def create_document_processing_tab(self):
        """
        创建文档处理标签页
        """
        tab = QWidget()
        self.tab_widget.addTab(tab, "文档处理")
        
        layout = QVBoxLayout(tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 文件上传组
        upload_group = QGroupBox("文件上传")
        upload_layout = QGridLayout(upload_group)
        
        # Excel文件选择
        upload_layout.addWidget(QLabel("Excel文件:"), 0, 0)
        self.excel_path_label = QLabel("未选择文件")
        self.excel_path_label.setStyleSheet("color: #666; font-style: italic;")
        upload_layout.addWidget(self.excel_path_label, 0, 1)
        
        self.browse_excel_btn = QPushButton("选择Excel文件")
        self.browse_excel_btn.clicked.connect(self.browse_excel_for_doc)
        upload_layout.addWidget(self.browse_excel_btn, 0, 2)
        
        # Word模板文件选择
        upload_layout.addWidget(QLabel("Word模板:"), 1, 0)
        self.word_path_label = QLabel("未选择文件")
        self.word_path_label.setStyleSheet("color: #666; font-style: italic;")
        upload_layout.addWidget(self.word_path_label, 1, 1)
        
        self.browse_word_btn = QPushButton("选择Word模板")
        self.browse_word_btn.clicked.connect(self.browse_word_template)
        upload_layout.addWidget(self.browse_word_btn, 1, 2)
        
        layout.addWidget(upload_group)
        
        # 处理选项组
        options_group = QGroupBox("处理选项")
        options_layout = QGridLayout(options_group)
        
        # 添加处理模式选择
        mode_label = QLabel("处理模式:")
        options_layout.addWidget(mode_label, 0, 0)
        
        # 创建单选按钮
        from PyQt5.QtWidgets import QRadioButton, QButtonGroup
        self.processing_mode_group = QButtonGroup()
        
        self.single_mode_radio = QRadioButton("单个处理")
        self.single_mode_radio.setToolTip("按公司名称分组，每个公司生成一个Word文档")
        self.single_mode_radio.setChecked(True)  # 默认选中
        self.processing_mode_group.addButton(self.single_mode_radio, 0)
        options_layout.addWidget(self.single_mode_radio, 0, 1)
        
        self.multiple_mode_radio = QRadioButton("多个处理")
        self.multiple_mode_radio.setToolTip("按群名称分组，同一群的数据生成一个Word文档")
        self.processing_mode_group.addButton(self.multiple_mode_radio, 1)
        options_layout.addWidget(self.multiple_mode_radio, 0, 2)
        
        # 添加模式说明
        mode_desc_label = QLabel("💡 单个处理：按公司名称分组，每个公司一个文档\n💡 多个处理：按群名称分组，同一群的数据合并到一个文档")
        mode_desc_label.setStyleSheet("color: #666; font-size: 12px; padding: 5px; background-color: #f9f9f9; border-radius: 3px;")
        options_layout.addWidget(mode_desc_label, 1, 0, 1, 3)
        
        options_layout.addWidget(QLabel("Excel工作表:"), 2, 0)
        self.sheet_selector = QComboBox()
        self.sheet_selector.setPlaceholderText("选择Excel工作表")
        self.sheet_selector.addItem("默认第一个工作表", None)
        options_layout.addWidget(self.sheet_selector, 2, 1)
        
        # 添加保存位置选择
        options_layout.addWidget(QLabel("保存位置:"), 3, 0)
        self.output_path_label = QLabel("未选择保存位置")
        self.output_path_label.setStyleSheet("color: #666; font-style: italic;")
        options_layout.addWidget(self.output_path_label, 3, 1)
        
        self.browse_output_btn = QPushButton("选择保存位置")
        self.browse_output_btn.clicked.connect(self.browse_output_location)
        options_layout.addWidget(self.browse_output_btn, 3, 2)
        
        layout.addWidget(options_group)
        
        # 处理按钮组
        process_group = QGroupBox("文档处理")
        process_layout = QVBoxLayout(process_group)
        
        # 按钮行
        button_layout = QHBoxLayout()
        
        self.process_doc_btn = QPushButton("开始处理文档")
        self.process_doc_btn.clicked.connect(self.process_documents)
        self.process_doc_btn.setEnabled(False)
        button_layout.addWidget(self.process_doc_btn)
        
        button_layout.addStretch()
        process_layout.addLayout(button_layout)
        
        # 进度条和状态
        progress_layout = QHBoxLayout()
        
        self.process_progress = QProgressBar()
        self.process_progress.setVisible(False)
        self.process_progress.setRange(0, 0)  # 不确定进度的进度条
        progress_layout.addWidget(self.process_progress)
        
        self.process_status_label = QLabel("")
        self.process_status_label.setStyleSheet("color: #666; font-style: italic;")
        progress_layout.addWidget(self.process_status_label)
        
        process_layout.addLayout(progress_layout)
        
        layout.addWidget(process_group)
        
        # 处理结果显示
        result_group = QGroupBox("处理结果")
        result_layout = QVBoxLayout(result_group)
        
        self.doc_result_text = QTextEdit()
        self.doc_result_text.setMaximumHeight(300)
        self.doc_result_text.setReadOnly(True)
        result_layout.addWidget(self.doc_result_text)
        
        # 快捷操作按钮
        button_layout = QHBoxLayout()
        
        self.open_folder_btn = QPushButton("📂 打开文件夹")
        self.open_folder_btn.setEnabled(False)
        self.open_folder_btn.clicked.connect(self.open_output_folder)
        button_layout.addWidget(self.open_folder_btn)
        
        button_layout.addStretch()
        result_layout.addLayout(button_layout)
        
        layout.addWidget(result_group)
        
        # 添加弹性空间
        layout.addStretch()
        
        # 初始化文档处理器
        self.document_processor = DocumentProcessor()
        self.processed_doc_path = None
    
    def browse_excel_for_doc(self):
        """
        为文档处理选择Excel文件
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_path_label.setText(os.path.basename(file_path))
            self.excel_path_label.setStyleSheet("color: #000;")
            self.excel_file_for_doc = file_path
            
            # 自动检测并填充工作表列表
            self.load_excel_sheets(file_path)
            
            self.check_files_ready()
    
    def load_excel_sheets(self, excel_path: str):
        """
        加载Excel文件的工作表列表
        """
        try:
            # 清空现有选项
            self.sheet_selector.clear()
            
            # 添加默认选项
            self.sheet_selector.addItem("默认第一个工作表", None)
            
            # 获取工作表列表
            sheets = self.document_processor.get_excel_sheets(excel_path)
            
            if sheets:
                # 添加所有工作表
                for sheet_name in sheets:
                    self.sheet_selector.addItem(f"📋 {sheet_name}", sheet_name)
                
                # 如果只有一个工作表，自动选择它
                if len(sheets) == 1:
                    self.sheet_selector.setCurrentIndex(1)  # 选择第一个实际工作表
                
                self.doc_result_text.append(f"✅ 检测到 {len(sheets)} 个工作表: {', '.join(sheets)}")
            else:
                self.doc_result_text.append("⚠️ 无法检测到工作表，将使用默认第一个工作表")
                
        except Exception as e:
            self.doc_result_text.append(f"❌ 检测工作表时出错: {str(e)}")
            # 保留默认选项
            pass
    
    def browse_word_template(self):
        """
        选择Word模板文件
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Word模板文件", "", "Word文档 (*.docx)"
        )
        
        if file_path:
            self.word_path_label.setText(os.path.basename(file_path))
            self.word_path_label.setStyleSheet("color: #000;")
            self.word_template_file = file_path
            self.check_files_ready()
    
    def browse_output_location(self):
        """
        选择输出文件保存位置
        """
        # 基于Word模板名称生成默认文件名
        if hasattr(self, 'word_template_file') and self.word_template_file:
            template_base_name = os.path.splitext(os.path.basename(self.word_template_file))[0]
            
            # 智能处理括号：如果模板名称已包含括号，则直接使用；否则添加括号
            import re
            if ('(' in template_base_name and ')' in template_base_name) or ('（' in template_base_name and '）' in template_base_name):
                # 如果已包含括号，替换括号内容为"公司名称"
                default_name = re.sub(r'[（(][^）)]*[）)]', '(公司名称)', template_base_name) + ".docx"
            else:
                # 如果不包含括号，添加括号和"公司名称"
                default_name = f"{template_base_name}(公司名称).docx"
        else:
            default_name = "VAT申报明细表(公司名称).docx"
            
        save_path, _ = QFileDialog.getSaveFileName(
            self, "选择保存位置", default_name, "Word文档 (*.docx)"
        )
        
        if save_path:
            # 显示文件夹路径而不是文件名
            folder_path = os.path.dirname(save_path)
            self.output_path_label.setText(folder_path)
            self.output_path_label.setStyleSheet("color: #000;")
            self.output_file_path = save_path
            self.check_files_ready()
    
    def check_files_ready(self):
        """
        检查文件是否都已选择，启用处理按钮
        """
        if (hasattr(self, 'excel_file_for_doc') and 
            hasattr(self, 'word_template_file') and
            hasattr(self, 'output_file_path')):
            self.process_doc_btn.setEnabled(True)
        else:
            self.process_doc_btn.setEnabled(False)
    
    def _get_file_size(self, file_path):
        """获取文件大小的友好显示格式"""
        try:
            if not file_path or not os.path.exists(file_path):
                return "未知"
            
            size_bytes = os.path.getsize(file_path)
            
            # 转换为合适的单位
            if size_bytes < 1024:
                return f"{size_bytes} B"
            elif size_bytes < 1024 * 1024:
                return f"{size_bytes / 1024:.1f} KB"
            elif size_bytes < 1024 * 1024 * 1024:
                return f"{size_bytes / (1024 * 1024):.1f} MB"
            else:
                return f"{size_bytes / (1024 * 1024 * 1024):.1f} GB"
        except Exception:
            return "未知"
    
    def open_output_folder(self):
        """打开输出文件夹"""
        try:
            if hasattr(self, 'last_output_path') and self.last_output_path:
                output_dir = os.path.dirname(self.last_output_path)
                if os.path.exists(output_dir):
                    # 根据操作系统打开文件夹
                    import subprocess
                    import platform
                    
                    system = platform.system()
                    if system == "Darwin":  # macOS
                        subprocess.run(["open", output_dir])
                    elif system == "Windows":
                        subprocess.run(["explorer", output_dir])
                    else:  # Linux
                        subprocess.run(["xdg-open", output_dir])
                else:
                    QMessageBox.warning(self, "警告", "输出文件夹不存在")
            else:
                QMessageBox.information(self, "提示", "没有可打开的文件夹")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"打开文件夹失败: {str(e)}")
    

    def process_documents(self):
        """
        处理文档 - 将Excel数据填充到Word表格
        """
        try:
            # 检查是否有正在处理的任务
            if hasattr(self, 'is_processing') and self.is_processing:
                QMessageBox.information(self, "提示", "文档正在处理中，请稍候...")
                return
            
            # 检查输出文件是否已存在，询问是否覆盖
            if os.path.exists(self.output_file_path):
                reply = QMessageBox.question(
                    self, "文件已存在", 
                    f"文件 {os.path.basename(self.output_file_path)} 已存在，是否覆盖？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply == QMessageBox.No:
                    return
            
            # 开始处理
            self.is_processing = True
            selected_sheet = self.sheet_selector.currentData()
            
            # 更新UI状态
            start_time = datetime.now()
            self.doc_result_text.clear()
            self.doc_result_text.append(f"🔄 开始处理文档... [{start_time.strftime('%Y-%m-%d %H:%M:%S')}]")
            self.process_doc_btn.setEnabled(False)
            self.process_doc_btn.setText("处理中...")
            
            # 显示状态
            self.process_status_label.setText("正在处理文档，请稍候...")
            
            # 使用用户选择的输出路径
            output_path = self.output_file_path
            
            # 获取字段映射
            column_mapping = create_default_column_mapping()
            
            self.doc_result_text.append(f"📊 使用工作表: {selected_sheet or '默认第一个工作表'}")
            self.process_status_label.setText("正在读取Excel数据...")
            
            # 获取处理模式
            processing_mode = "single" if self.single_mode_radio.isChecked() else "multiple"
            self.doc_result_text.append(f"🔧 处理模式: {'单个处理（按公司名称分组）' if processing_mode == 'single' else '多个处理（按群名称分组）'}")
            
            # 执行文档处理
            result = self.document_processor.process_documents(
                excel_path=self.excel_file_for_doc,
                word_template_path=self.word_template_file,
                output_path=output_path,
                sheet_name=selected_sheet,
                column_mapping=column_mapping,
                processing_mode=processing_mode
            )
            
            if result['success']:
                end_time = datetime.now()
                processing_time = (end_time - start_time).total_seconds()
                
                self.doc_result_text.append(f"✅ 文档处理成功！ [{end_time.strftime('%Y-%m-%d %H:%M:%S')}]")
                self.doc_result_text.append(f"⏱️ 处理耗时: {processing_time:.2f} 秒")
                self.doc_result_text.append(f"📊 数据统计:")
                self.doc_result_text.append(f"   • 总数据行数: {result.get('total_rows_detected', 0)}")
                self.doc_result_text.append(f"   • 已处理行数: {result['rows_filled']}")
                
                # 根据处理模式显示不同的详细信息
                if processing_mode == "single":
                    self.doc_result_text.append(f"📄 处理模式: 单个文档处理")
                    self.doc_result_text.append(f"💾 生成文件: {os.path.basename(result['output_path'])}")
                    file_size = self._get_file_size(result['output_path'])
                    self.doc_result_text.append(f"📏 文件大小: {file_size}")
                    
                elif processing_mode == "company":
                    self.doc_result_text.append(f"📄 处理模式: 按公司分组处理")
                    self.doc_result_text.append(f"🏢 生成公司文档数: {result.get('groups_count', 0)}")
                    self.doc_result_text.append(f"📁 生成的文件:")
                    for i, file_path in enumerate(result.get('generated_files', []), 1):
                        file_size = self._get_file_size(file_path)
                        self.doc_result_text.append(f"   {i}. {os.path.basename(file_path)} ({file_size})")
                        
                elif processing_mode == "group":
                    self.doc_result_text.append(f"📄 处理模式: 按群组分组处理")
                    self.doc_result_text.append(f"👥 生成群组文档数: {result.get('groups_count', 0)}")
                    self.doc_result_text.append(f"📁 生成的文件:")
                    for i, file_path in enumerate(result.get('generated_files', []), 1):
                        file_size = self._get_file_size(file_path)
                        self.doc_result_text.append(f"   {i}. {os.path.basename(file_path)} ({file_size})")
                
                # 显示文件位置
                output_dir = os.path.dirname(result['output_path']) if result['output_path'] else ""
                self.doc_result_text.append(f"📂 文件位置: {output_dir}")
                self.doc_result_text.append("=" * 50)
                
                self.process_status_label.setText("处理完成！文档已保存")
                
                # 保存输出路径并启用快捷按钮
                self.last_output_path = result['output_path']
                self.open_folder_btn.setEnabled(True)
                
                # 状态栏显示简要信息
                if processing_mode == "single":
                    self.status_bar.showMessage(f"文档已保存: {os.path.basename(result['output_path'])}")
                else:
                    file_count = len(result.get('generated_files', []))
                    self.status_bar.showMessage(f"已生成 {file_count} 个文档，处理 {result['rows_filled']} 行数据")
            else:
                error_time = datetime.now()
                self.doc_result_text.append(f"❌ 处理失败: {result['error']} [{error_time.strftime('%Y-%m-%d %H:%M:%S')}]")
                self.process_status_label.setText("处理失败")
                self.status_bar.showMessage("文档处理失败")
                
                # 禁用快捷按钮
                self.open_folder_btn.setEnabled(False)
                
        except ValueError as e:
            error_time = datetime.now()
            self.doc_result_text.append(f"❌ 参数错误: {str(e)} [{error_time.strftime('%Y-%m-%d %H:%M:%S')}]")
            QMessageBox.warning(self, "参数错误", f"请输入有效的数字: {str(e)}")
            self.process_status_label.setText("参数错误")
            # 禁用快捷按钮
            self.open_folder_btn.setEnabled(False)
        except Exception as e:
            error_time = datetime.now()
            self.doc_result_text.append(f"❌ 处理过程中发生错误: {str(e)} [{error_time.strftime('%Y-%m-%d %H:%M:%S')}]")
            QMessageBox.critical(self, "处理错误", f"文档处理失败: {str(e)}")
            self.process_status_label.setText("处理出错")
            # 禁用快捷按钮
            self.open_folder_btn.setEnabled(False)
        finally:
            # 恢复UI状态
            self.is_processing = False
            self.process_doc_btn.setEnabled(True)
            self.process_doc_btn.setText("开始处理文档")
            self.process_status_label.setText("")
    



def main():
    try:
        app = QApplication(sys.argv)
        app.setApplicationName("VAT验证工具")
        app.setApplicationVersion("2.0")
        
        # 设置应用图标（如果有的话）
        # app.setWindowIcon(QIcon('icon.png'))
        
        print("正在初始化VAT验证工具...")
        
        # 检查关键模块是否可用
        try:
            from document_processor import DocumentProcessor, create_default_column_mapping
            print("✓ document_processor模块加载成功")
        except ImportError as e:
            print(f"✗ document_processor模块加载失败: {e}")
            QMessageBox.critical(None, "模块加载错误", f"无法加载document_processor模块: {e}")
            return
        
        try:
            import openpyxl
            print("✓ openpyxl模块加载成功")
        except ImportError as e:
            print(f"✗ openpyxl模块加载失败: {e}")
            QMessageBox.critical(None, "模块加载错误", f"无法加载openpyxl模块: {e}")
            return
        
        window = VATValidatorGUI()
        print("✓ 主窗口创建成功")
        
        window.show()
        print("✓ 应用程序启动成功")
        
        sys.exit(app.exec_())
        
    except Exception as e:
        print(f"✗ 应用程序启动失败: {e}")
        import traceback
        traceback.print_exc()
        
        # 尝试显示错误对话框
        try:
            if 'app' not in locals():
                app = QApplication(sys.argv)
            QMessageBox.critical(None, "启动错误", f"应用程序启动失败:\n{str(e)}\n\n请检查控制台输出获取详细信息。")
        except:
            pass

if __name__ == "__main__":
    main()