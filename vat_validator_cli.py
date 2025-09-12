#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VAT验证工具 - 命令行版本
支持Excel文件导入和结果导出
"""

import requests
from openpyxl import load_workbook, Workbook
from datetime import datetime
import json
import os
import sys

class VATValidatorCLI:
    def __init__(self):
        self.batch_results = []
        
    def extract_country_code(self, vat_number):
        """
        从VAT号码中提取国家代码
        """
        # 常见的欧盟国家代码
        eu_countries = {
            'AT', 'BE', 'BG', 'CY', 'CZ', 'DE', 'DK', 'EE', 'ES', 'FI',
            'FR', 'GR', 'HR', 'HU', 'IE', 'IT', 'LT', 'LU', 'LV', 'MT',
            'NL', 'PL', 'PT', 'RO', 'SE', 'SI', 'SK'
        }
        
        # 检查前两位是否为国家代码
        if len(vat_number) >= 2:
            potential_code = vat_number[:2].upper()
            if potential_code in eu_countries:
                return potential_code
        
        return None
    
    def verify_vat(self, country_code, vat_number):
        """
        验证单个VAT号码
        """
        try:
            # 构建API URL
            url = f"https://ec.europa.eu/taxation_customs/vies/rest-api/ms/{country_code}/vat/{vat_number}"
            
            # 设置请求头
            headers = {
                'User-Agent': 'VAT-Validator/1.0',
                'Accept': 'application/json'
            }
            
            # 发送请求
            response = requests.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                return {
                    'success': True,
                    'vat_number': f"{country_code}{vat_number}",
                    'country_code': country_code,
                    'data': data
                }
            else:
                return {
                    'success': False,
                    'vat_number': f"{country_code}{vat_number}",
                    'country_code': country_code,
                    'error': f"HTTP错误: {response.status_code}"
                }
                
        except requests.exceptions.RequestException as e:
            return {
                'success': False,
                'vat_number': f"{country_code}{vat_number}",
                'country_code': country_code,
                'error': f"网络错误: {str(e)}"
            }
        except Exception as e:
            return {
                'success': False,
                'vat_number': f"{country_code}{vat_number}",
                'country_code': country_code,
                'error': f"未知错误: {str(e)}"
            }
    
    def import_excel(self, file_path):
        """
        从Excel文件导入VAT号码
        """
        try:
            print(f"正在读取Excel文件: {file_path}")
            
            # 读取Excel文件
            wb = load_workbook(file_path)
            ws = wb.active
            
            if ws is None:
                print("错误: 无法读取Excel工作表")
                return False
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value is not None else '')
            
            print(f"发现列: {headers}")
            
            # 检查必要的列
            required_columns = ['NIF Contraparte', 'Importe', 'Tipo']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                print(f"错误: Excel文件缺少必要的列: {', '.join(missing_columns)}")
                return False
            
            # 找到NIF Contraparte列的索引
            nif_col_index = None
            for i, header in enumerate(headers):
                if header == 'NIF Contraparte':
                    nif_col_index = i + 1  # openpyxl使用1基索引
                    break
            
            if nif_col_index is None:
                print("错误: 未找到NIF Contraparte列")
                return False
            
            # 提取VAT号码和原始数据
            vat_data = []
            if ws is not None:
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if row and len(row) >= nif_col_index and row[nif_col_index - 1]:
                        vat_number = str(row[nif_col_index - 1]).strip()
                        if vat_number:
                            row_data = {}
                            for i in range(min(len(headers), len(row))):
                                row_data[headers[i]] = row[i]
                            row_data['_row_number'] = row_num
                            vat_data.append((vat_number, row_data))
            
            if not vat_data:
                print("错误: 未找到有效的VAT号码")
                return False
            
            print(f"成功导入 {len(vat_data)} 个VAT号码")
            
            # 开始批量验证
            self.start_batch_verification(vat_data)
            return True
            
        except Exception as e:
            print(f"读取Excel文件失败: {str(e)}")
            return False
    
    def start_batch_verification(self, vat_data):
        """
        开始批量验证
        """
        self.batch_results = []
        total = len(vat_data)
        
        print(f"\n开始验证 {total} 个VAT号码...")
        print("=" * 60)
        
        for i, (vat_number, original_data) in enumerate(vat_data, 1):
            print(f"\n[{i}/{total}] 正在验证: {vat_number}")
            
            # 尝试从VAT号码中提取国家代码
            country_code = self.extract_country_code(vat_number)
            
            if country_code:
                # 移除国家代码前缀
                clean_vat = vat_number[2:] if vat_number.startswith(country_code) else vat_number
                result = self.verify_vat(country_code, clean_vat)
            else:
                # 如果无法提取国家代码，尝试使用意大利作为默认值
                result = self.verify_vat('IT', vat_number)
            
            # 添加原始数据
            result['original_data'] = original_data
            
            # 显示结果
            if result['success']:
                data = result['data']
                is_valid = data.get('isValid', False)
                company_name = data.get('name', '未提供')
                print(f"✅ {result['vat_number']}: {'有效' if is_valid else '无效'}")
                if is_valid:
                    print(f"   公司名称: {company_name}")
            else:
                print(f"❌ {result['vat_number']}: 验证失败 - {result['error']}")
            
            self.batch_results.append(result)
        
        print("\n=" * 60)
        print(f"批量验证完成，共验证 {total} 个VAT号码")
        
        # 统计结果
        successful = sum(1 for r in self.batch_results if r['success'])
        valid = sum(1 for r in self.batch_results if r['success'] and r.get('data', {}).get('isValid', False))
        
        print(f"验证成功: {successful}/{total}")
        print(f"有效VAT: {valid}/{total}")
    
    def export_results(self, file_path):
        """
        导出验证结果到Excel文件
        """
        if not self.batch_results:
            print("没有验证结果可导出")
            return False
        
        try:
            print(f"\n正在导出结果到: {file_path}")
            
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "VAT验证结果"
            
            # 设置表头
            headers = ['VAT号码', '国家代码', '验证状态', '是否有效', '公司名称', '地址', '请求日期', '错误信息']
            
            # 添加原始数据的列头
            if self.batch_results and 'original_data' in self.batch_results[0]:
                original_headers = list(self.batch_results[0]['original_data'].keys())
                headers.extend([f'原始_{header}' for header in original_headers if header != '_row_number'])
            
            # 写入表头
            if ws is not None:
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 写入数据
                for row_idx, result in enumerate(self.batch_results, 2):
                    ws.cell(row=row_idx, column=1, value=result['vat_number'])
                    ws.cell(row=row_idx, column=2, value=result['country_code'])
                    ws.cell(row=row_idx, column=3, value='成功' if result['success'] else '失败')
                    
                    if result['success'] and 'data' in result:
                        data = result['data']
                        ws.cell(row=row_idx, column=4, value='是' if data.get('isValid', False) else '否')
                        ws.cell(row=row_idx, column=5, value=data.get('name', ''))
                        ws.cell(row=row_idx, column=6, value=data.get('address', ''))
                        ws.cell(row=row_idx, column=7, value=data.get('requestDate', ''))
                    else:
                        ws.cell(row=row_idx, column=8, value=result.get('error', ''))
                    
                    # 添加原始数据
                    if 'original_data' in result:
                        col_idx = 9
                        for key, value in result['original_data'].items():
                            if key != '_row_number':
                                ws.cell(row=row_idx, column=col_idx, value=value)
                                col_idx += 1
            
            # 保存文件
            wb.save(file_path)
            print(f"✅ 结果已成功导出到: {file_path}")
            return True
            
        except Exception as e:
            print(f"导出失败: {str(e)}")
            return False

def main():
    """
    主函数
    """
    print("VAT验证工具 - 命令行版本")
    print("=" * 40)
    
    validator = VATValidatorCLI()
    
    while True:
        print("\n请选择操作:")
        print("1. 导入Excel文件进行批量验证")
        print("2. 单个VAT号码验证")
        print("3. 导出验证结果")
        print("4. 退出")
        
        choice = input("\n请输入选择 (1-4): ").strip()
        
        if choice == '1':
            file_path = input("请输入Excel文件路径: ").strip()
            if os.path.exists(file_path):
                validator.import_excel(file_path)
            else:
                print("文件不存在，请检查路径")
        
        elif choice == '2':
            vat_number = input("请输入VAT号码: ").strip()
            if vat_number:
                # 尝试提取国家代码
                country_code = validator.extract_country_code(vat_number)
                if country_code:
                    clean_vat = vat_number[2:] if vat_number.startswith(country_code) else vat_number
                else:
                    country_code = input("请输入国家代码 (如IT, DE, FR等): ").strip().upper()
                    clean_vat = vat_number
                
                print(f"\n正在验证: {country_code}{clean_vat}")
                result = validator.verify_vat(country_code, clean_vat)
                
                if result['success']:
                    data = result['data']
                    print(f"✅ 验证成功")
                    print(f"VAT号码: {result['vat_number']}")
                    print(f"是否有效: {'是' if data.get('isValid', False) else '否'}")
                    print(f"公司名称: {data.get('name', '未提供')}")
                    print(f"地址: {data.get('address', '未提供')}")
                else:
                    print(f"❌ 验证失败: {result['error']}")
        
        elif choice == '3':
            if validator.batch_results:
                output_path = input("请输入导出文件路径 (如: results.xlsx): ").strip()
                if not output_path:
                    output_path = f"vat_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                validator.export_results(output_path)
            else:
                print("没有验证结果可导出")
        
        elif choice == '4':
            print("再见！")
            break
        
        else:
            print("无效选择，请重新输入")

if __name__ == "__main__":
    main()