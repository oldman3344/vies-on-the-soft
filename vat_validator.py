#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VIES VAT验证工具
复制欧盟VIES网站的VAT号码验证功能
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import requests
from openpyxl import load_workbook, Workbook
from datetime import datetime
import json
import threading
import os

class VATValidator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("VIES VAT验证工具")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # 欧盟成员国代码
        self.country_codes = {
            "奥地利 (AT)": "AT",
            "比利时 (BE)": "BE",
            "保加利亚 (BG)": "BG",
            "克罗地亚 (HR)": "HR",
            "塞浦路斯 (CY)": "CY",
            "捷克 (CZ)": "CZ",
            "丹麦 (DK)": "DK",
            "爱沙尼亚 (EE)": "EE",
            "芬兰 (FI)": "FI",
            "法国 (FR)": "FR",
            "德国 (DE)": "DE",
            "希腊 (EL)": "EL",
            "匈牙利 (HU)": "HU",
            "爱尔兰 (IE)": "IE",
            "意大利 (IT)": "IT",
            "拉脱维亚 (LV)": "LV",
            "立陶宛 (LT)": "LT",
            "卢森堡 (LU)": "LU",
            "马耳他 (MT)": "MT",
            "荷兰 (NL)": "NL",
            "波兰 (PL)": "PL",
            "葡萄牙 (PT)": "PT",
            "罗马尼亚 (RO)": "RO",
            "斯洛伐克 (SK)": "SK",
            "斯洛文尼亚 (SI)": "SI",
            "西班牙 (ES)": "ES",
            "瑞典 (SE)": "SE",
            "北爱尔兰 (XI)": "XI"
        }
        
        self.setup_ui()
        
    def setup_ui(self):
        """设置用户界面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky='wens')
        
        # 配置网格权重
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # 标题
        title_label = ttk.Label(main_frame, text="VIES VAT号码验证工具", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # 单个验证区域
        single_frame = ttk.LabelFrame(main_frame, text="单个VAT号码验证", padding="10")
        single_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        single_frame.columnconfigure(1, weight=1)
        
        # 国家选择
        ttk.Label(single_frame, text="成员国/北爱尔兰:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.country_var = tk.StringVar()
        self.country_combo = ttk.Combobox(single_frame, textvariable=self.country_var, 
                                         values=list(self.country_codes.keys()), 
                                         state="readonly", width=30)
        self.country_combo.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        self.country_combo.set("意大利 (IT)")  # 默认选择意大利
        
        # VAT号码输入
        ttk.Label(single_frame, text="VAT号码:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        self.vat_var = tk.StringVar()
        self.vat_entry = ttk.Entry(single_frame, textvariable=self.vat_var, width=30)
        self.vat_entry.grid(row=1, column=1, sticky="ew", padx=(0, 10), pady=(10, 0))
        
        # 验证按钮
        self.verify_btn = ttk.Button(single_frame, text="验证", command=self.verify_single_vat)
        self.verify_btn.grid(row=1, column=2, pady=(10, 0))
        
        # 结果显示
        self.result_text = tk.Text(single_frame, height=6, width=80)
        self.result_text.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        
        # 滚动条
        scrollbar = ttk.Scrollbar(single_frame, orient="vertical", command=self.result_text.yview)
        scrollbar.grid(row=2, column=3, sticky="ns", pady=(10, 0))
        self.result_text.configure(yscrollcommand=scrollbar.set)
        
        # 批量验证区域
        batch_frame = ttk.LabelFrame(main_frame, text="批量VAT号码验证", padding="10")
        batch_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        batch_frame.columnconfigure(1, weight=1)
        
        # 文件操作按钮
        ttk.Button(batch_frame, text="导入Excel文件", command=self.import_excel).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(batch_frame, text="导出结果", command=self.export_results).grid(row=0, column=1, padx=(0, 10))
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(batch_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        
        # 状态标签
        self.status_var = tk.StringVar(value="就绪")
        self.status_label = ttk.Label(batch_frame, textvariable=self.status_var)
        self.status_label.grid(row=2, column=0, columnspan=2, pady=(5, 0))
        
        # 存储批量验证结果
        self.batch_results = []
        
    def verify_vat(self, country_code, vat_number):
        """验证单个VAT号码"""
        try:
            # 构建API URL
            url = f"https://ec.europa.eu/taxation_customs/vies/rest-api/ms/{country_code}/vat/{vat_number}"
            
            # 设置请求头
            headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-encoding': 'gzip, deflate, br, zstd',
                'accept-language': 'zh-CN,zh;q=0.9',
                'cache-control': 'No-Cache',
                'connection': 'keep-alive',
                'host': 'ec.europa.eu',
                'pragma': 'no-cache',
                'referer': 'https://ec.europa.eu/taxation_customs/vies/',
                'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"macOS"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-origin',
                'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
                'x-requested-with': 'XMLHttpRequest'
            }
            
            # 发送请求
            response = requests.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                return {
                    'success': True,
                    'data': data,
                    'country_code': country_code,
                    'vat_number': vat_number
                }
            else:
                return {
                    'success': False,
                    'error': f"HTTP错误: {response.status_code}",
                    'country_code': country_code,
                    'vat_number': vat_number
                }
                
        except requests.exceptions.RequestException as e:
            return {
                'success': False,
                'error': f"网络错误: {str(e)}",
                'country_code': country_code,
                'vat_number': vat_number
            }
        except Exception as e:
            return {
                'success': False,
                'error': f"未知错误: {str(e)}",
                'country_code': country_code,
                'vat_number': vat_number
            }
    
    def verify_single_vat(self):
        """验证单个VAT号码"""
        country_name = self.country_var.get()
        vat_number = self.vat_var.get().strip()
        
        if not country_name or not vat_number:
            messagebox.showerror("错误", "请选择国家并输入VAT号码")
            return
        
        country_code = self.country_codes[country_name]
        
        # 清空结果显示
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, "正在验证...\n")
        self.verify_btn.config(state="disabled")
        
        def verify_thread():
            result = self.verify_vat(country_code, vat_number)
            
            # 在主线程中更新UI
            self.root.after(0, lambda: self.display_single_result(result))
        
        # 在新线程中执行验证
        threading.Thread(target=verify_thread, daemon=True).start()
    
    def display_single_result(self, result):
        """显示单个验证结果"""
        self.result_text.delete(1.0, tk.END)
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.result_text.insert(tk.END, f"验证时间: {timestamp}\n")
        self.result_text.insert(tk.END, f"国家代码: {result['country_code']}\n")
        self.result_text.insert(tk.END, f"VAT号码: {result['vat_number']}\n")
        self.result_text.insert(tk.END, "-" * 50 + "\n")
        
        if result['success']:
            data = result['data']
            self.result_text.insert(tk.END, f"验证状态: {'有效' if data.get('valid', False) else '无效'}\n")
            
            if 'name' in data and data['name']:
                self.result_text.insert(tk.END, f"公司名称: {data['name']}\n")
            
            if 'address' in data and data['address']:
                self.result_text.insert(tk.END, f"地址: {data['address']}\n")
            
            if 'requestDate' in data:
                self.result_text.insert(tk.END, f"请求日期: {data['requestDate']}\n")
                
            # 显示完整的JSON响应
            self.result_text.insert(tk.END, "\n完整响应:\n")
            self.result_text.insert(tk.END, json.dumps(data, indent=2, ensure_ascii=False))
        else:
            self.result_text.insert(tk.END, f"验证失败: {result['error']}\n")
        
        self.verify_btn.config(state="normal")
    
    def import_excel(self):
        """导入Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        try:
            # 读取Excel文件
            wb = load_workbook(file_path)
            ws = wb.active
            
            if ws is None:
                messagebox.showerror("错误", "无法读取Excel工作表")
                return
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value is not None else '')
            
            # 检查必要的列
            required_columns = ['NIF Contraparte', 'Importe', 'Tipo']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                messagebox.showerror("错误", f"Excel文件缺少必要的列: {', '.join(missing_columns)}")
                return
            
            # 找到NIF Contraparte列的索引
            nif_col_index = None
            for i, header in enumerate(headers):
                if header == 'NIF Contraparte':
                    nif_col_index = i + 1  # openpyxl使用1基索引
                    break
            
            if nif_col_index is None:
                messagebox.showerror("错误", "未找到NIF Contraparte列")
                return
            
            # 提取VAT号码和原始数据
            vat_data = []
            if ws is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= nif_col_index and row[nif_col_index - 1]:  # 检查VAT号码不为空
                        vat_number = str(row[nif_col_index - 1]).strip()
                        if vat_number:
                            row_data = {}
                            for i in range(min(len(headers), len(row))):
                                row_data[headers[i]] = row[i]
                            vat_data.append((vat_number, row_data))
            
            if not vat_data:
                messagebox.showerror("错误", "未找到有效的VAT号码")
                return
            
            # 开始批量验证
            self.start_batch_verification(vat_data)
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
    
    def start_batch_verification(self, vat_data):
        """开始批量验证"""
        self.batch_results = []
        self.progress_var.set(0)
        
        def verify_batch():
            total = len(vat_data)
            
            for i, (vat_number, original_data) in enumerate(vat_data):
                # 更新状态
                self.root.after(0, lambda i=i, total=total, vat=vat_number: 
                    self.status_var.set(f"正在验证 {i+1}/{total}: {vat}"))
                
                # 尝试从VAT号码中提取国家代码
                country_code = self.extract_country_code(vat_number)
                
                if country_code:
                    # 移除国家代码前缀
                    clean_vat = vat_number[2:] if vat_number.startswith(country_code) else vat_number
                    result = self.verify_vat(country_code, clean_vat)
                else:
                    # 如果无法提取国家代码，尝试使用意大利作为默认值
                    result = self.verify_vat('IT', vat_number)
                
                # 添加原始数据
                result['original_data'] = original_data
                
                self.batch_results.append(result)
                
                # 更新进度条
                progress = ((i + 1) / total) * 100
                self.root.after(0, lambda p=progress: self.progress_var.set(p))
            
            # 完成
            self.root.after(0, lambda: self.status_var.set(f"批量验证完成，共验证 {total} 个VAT号码"))
        
        # 在新线程中执行批量验证
        threading.Thread(target=verify_batch, daemon=True).start()
    
    def extract_country_code(self, vat_number):
        """从VAT号码中提取国家代码"""
        vat_number = vat_number.upper().strip()
        
        # 检查是否以已知的国家代码开头
        for country_code in self.country_codes.values():
            if vat_number.startswith(country_code):
                return country_code
        
        return None
    
    def export_results(self):
        """导出验证结果"""
        if not self.batch_results:
            messagebox.showwarning("警告", "没有可导出的结果")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存结果文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        
        if not file_path:
            return
        
        try:
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "VAT验证结果"
            
            # 设置表头
            headers = ['VAT号码', '国家代码', '验证状态', '是否有效', '公司名称', '地址', '请求日期', '错误信息']
            
            # 添加原始数据的列头
            if self.batch_results and 'original_data' in self.batch_results[0]:
                original_headers = list(self.batch_results[0]['original_data'].keys())
                headers.extend([f'原始_{header}' for header in original_headers])
            
            # 写入表头
            if ws is not None:
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 写入数据
                for row_idx, result in enumerate(self.batch_results, 2):
                    ws.cell(row=row_idx, column=1, value=result['vat_number'])
                    ws.cell(row=row_idx, column=2, value=result['country_code'])
                    ws.cell(row=row_idx, column=3, value='成功' if result['success'] else '失败')
                    
                    if result['success'] and 'data' in result:
                        data = result['data']
                        ws.cell(row=row_idx, column=4, value='是' if data.get('valid', False) else '否')
                        ws.cell(row=row_idx, column=5, value=data.get('name', ''))
                        ws.cell(row=row_idx, column=6, value=data.get('address', ''))
                        ws.cell(row=row_idx, column=7, value=data.get('requestDate', ''))
                    else:
                        ws.cell(row=row_idx, column=8, value=result.get('error', ''))
                    
                    # 添加原始数据
                    if 'original_data' in result:
                        for col_idx, (key, value) in enumerate(result['original_data'].items(), 9):
                            ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 保存文件
            wb.save(file_path)
            
            messagebox.showinfo("成功", f"结果已导出到: {file_path}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def run(self):
        """运行应用程序"""
        self.root.mainloop()

if __name__ == "__main__":
    app = VATValidator()
    app.run()